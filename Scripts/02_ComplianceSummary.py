"""02 Compliance Summary: fill template counts, build Appendix B/C, export PDFs."""

import argparse
import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font

from env import (
    ALIGN_CENTER_WRAP,
    ALIGN_LEFT_WRAP,
    FONT_NAME,
    GRAY_FILL,
    LINK_FONT_11,
    TEMPLATE_DIR,
    THIN_BORDER,
    apply_print_setup,
    clean_text,
    export_workbook_to_pdf,
    fetch_xml,
    output_subdir,
    parse_results,
    save_workbook,
    script_dir,
    xml_cache_path,
    xml_source_url,
)

URL = xml_source_url("compliance_summary")
CACHE_FILE = xml_cache_path("compliance_summary")
OUT_DIR = output_subdir("02_ComplianceSummary")
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, "Requirement_Compliance_Summary.xlsx")
SUMMARY_FILE = os.path.join(OUT_DIR, "Requirement_Compliance_Summary.xlsx")
APPENDIX_B_FILE = os.path.join(
    OUT_DIR, "Appendix B – List Of Changes Agreed By MTR In Artifact Types.xlsx"
)
APPENDIX_C_FILE = os.path.join(
    OUT_DIR, "Appendix C – List Of Proposed Changes In Artifact Types.xlsx"
)

# NAME_ALIASES = {
#     "1263 - Contract 12051.docx": "1263 - Contract 12501.docx",
# }

INVALID_SHEET_CHARS = re.compile(r'[:\\/?*\[\]]')


def normalize_module_name(name):
    name = (name or "").strip()
    name = re.sub(r"^\(Main Scope\)\s*", "", name)
    # name = NAME_ALIASES.get(name, name)
    return name


def empty_counts():
    return {
        "info": 0,  # F
        "tv_req": 0,  # G
        "tv_init_no": 0,  # H
        "compliant": 0,  # I
        "partial": 0,  # J
        "not_compliant": 0,  # K
        "tv_yes_planned": 0,  # L / O
        "tv_no_na": 0,  # M
        "blank_dcs": 0,  # N
    }


def is_requirement(artifact_type):
    return "requirement" in (artifact_type or "").lower()


def count_row(row, counts):
    artifact = (row.get("REQUIREMENT_TYPE") or "").strip().lower()
    tv = (row.get("LITERAL_NAME") or "").strip().lower()
    init = (row.get("LITERAL_NAME1") or "").strip().lower()
    dcs = (row.get("LITERAL_NAME2") or "").strip().lower()

    if artifact == "information":
        counts["info"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes":
        counts["tv_req"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and init == "no" and dcs == "compliant":
        counts["tv_init_no"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and dcs == "compliant":
        counts["compliant"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and dcs == "partially compliant":
        counts["partial"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and dcs == "not compliant":
        counts["not_compliant"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and dcs == "planned (not started)":
        counts["tv_yes_planned"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "no" and dcs == "n/a":
        counts["tv_no_na"] += 1

    if is_requirement(row.get("REQUIREMENT_TYPE") or "") and tv == "yes" and dcs == "":
        counts["blank_dcs"] += 1


def aggregate_counts(rows):
    by_name = defaultdict(empty_counts)
    seen_names = defaultdict(int)
    for row in rows:
        name = normalize_module_name(row.get("NAME", ""))
        seen_names[name] += 1
        count_row(row, by_name[name])
    return by_name, seen_names


def normalized_header_text(value):
    text = clean_text(str(value or ""))
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def locate_summary_layout(ws):
    """Detect summary count and flag columns from the template headers."""
    app_b_col = None
    app_c_col = None
    count_start_col = None
    max_col = ws.max_column or 1

    for col in range(1, max_col + 1):
        row1 = normalized_header_text(ws.cell(row=1, column=col).value)
        row3 = normalized_header_text(ws.cell(row=3, column=col).value)
        combined = " ".join(part for part in (row1, row3) if part)

        if "include in app b" in combined:
            app_b_col = col
        elif "include in app c" in combined:
            app_c_col = col

        if count_start_col is None and "(1)" in combined and "information artifact" in combined:
            count_start_col = col

    missing = []
    if app_b_col is None:
        missing.append("App B flag column")
    if app_c_col is None:
        missing.append("App C flag column")
    if count_start_col is None:
        missing.append("count start column")
    if missing:
        raise SystemExit(f"Could not detect summary layout: {', '.join(missing)}")

    return {
        "app_b_col": app_b_col,
        "app_c_col": app_c_col,
        "count_start_col": count_start_col,
    }


def fill_summary(ws, by_name, layout):
    """Write the 10 count columns without overwriting the App B/C flags."""
    font = Font(name=FONT_NAME, size=11)
    last_row = ws.max_row or 5
    catalog_names = set()
    count_cols = {
        layout["count_start_col"] + 0: "info",
        layout["count_start_col"] + 1: "tv_req",
        layout["count_start_col"] + 2: "tv_init_no",
        layout["count_start_col"] + 3: "compliant",
        layout["count_start_col"] + 4: "partial",
        layout["count_start_col"] + 5: "not_compliant",
        layout["count_start_col"] + 6: "tv_yes_planned",
        layout["count_start_col"] + 7: "tv_no_na",
        layout["count_start_col"] + 8: "blank_dcs",
        layout["count_start_col"] + 9: "tv_yes_planned",
    }

    for excel_row in range(5, last_row + 1):
        name = (ws.cell(row=excel_row, column=1).value or "").strip()
        header = (ws.cell(row=excel_row, column=2).value or "").strip()
        if not header:
            continue
        if not name:
            continue

        catalog_names.add(name)
        counts = by_name.get(name, empty_counts())
        for col, key in count_cols.items():
            val = counts[key]
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = font
            cell.alignment = ALIGN_CENTER_WRAP
            cell.border = THIN_BORDER

    return catalog_names


def clean_sheet_name(name):
    name = INVALID_SHEET_CHARS.sub("", name or "")
    name = name.strip()
    if len(name) > 31:
        name = name[:31]
    return name or "Sheet1"


def unique_sheet_name(wb, base_name):
    candidate = base_name
    counter = 1
    existing = {ws.title.lower() for ws in wb.worksheets}
    while candidate.lower() in existing:
        suffix = f"_{counter}"
        candidate = base_name[: 31 - len(suffix)] + suffix
        counter += 1
        if counter > 1000:
            break
    return candidate


def include_flag(value):
    if value is True:
        return True
    if isinstance(value, (int, float)) and value == 1:
        return True
    text = normalized_header_text(value)
    return text in {"true", "yes", "y", "1"}


def appendix_modules(ws, appendix_type, layout):
    """Return ordered list of (module_name, sheet_title) for App B or App C."""
    flag_col = layout["app_b_col"] if appendix_type == "B" else layout["app_c_col"]
    modules = []
    seen = set()
    last_row = ws.max_row or 5
    for excel_row in range(5, last_row + 1):
        name = (ws.cell(row=excel_row, column=1).value or "").strip()
        header = (ws.cell(row=excel_row, column=2).value or "").strip()
        if not header:
            continue
        if not name:
            continue
        flag = ws.cell(row=excel_row, column=flag_col).value
        if not include_flag(flag):
            continue
        if name in seen:
            continue
        seen.add(name)
        modules.append((name, clean_sheet_name(header)))
    return modules


def filter_appendix_rows(rows, module_name, filter_keyword):
    filtered = []
    for row in rows:
        name = normalize_module_name(row.get("NAME", ""))
        if name != module_name:
            continue
        action = (row.get("LITERAL_NAME3") or "").lower()
        status = (row.get("LITERAL_NAME4") or "").strip().lower()
        if "info" not in action and "heading" not in action:
            continue
        if status != filter_keyword.lower():
            continue
        filtered.append(row)
    return filtered


def change_in_artifact_type(action):
    action_l = (action or "").lower()
    if "info" in action_l:
        return "Information"
    if "heading" in action_l:
        return "Heading"
    return ""


def format_appendix_sheet(ws, last_row, left_header):
    ws.column_dimensions["A"].width = 8.11
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    font_regular = Font(name=FONT_NAME, size=11)
    font_header = Font(name=FONT_NAME, size=11, bold=True)

    for r in range(1, last_row + 1):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER_WRAP
            if r == 1:
                cell.font = font_header
                cell.fill = GRAY_FILL
            else:
                cell.font = font_regular

    if last_row >= 2:
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=2)
            cell.alignment = ALIGN_LEFT_WRAP
            cell.font = LINK_FONT_11

    apply_print_setup(
        ws,
        left_header=f'&"Arial,Regular"&11 {left_header}',
        right_header='&"Arial,Regular"&11 &[Tab]',
        orientation="landscape",
        fit_width=1,
        print_title_rows="1:1",
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
    )
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    ws.oddFooter.right.text = '&"Arial,Regular"&11 Page &[Page] of &[Pages]'
    ws.oddFooter.right.size = 11


def write_appendix_workbook(
    modules, rows, filter_keyword, appendix_type, output_path, left_header
):
    sheets_data = []
    for module_name, sheet_base in modules:
        filtered = filter_appendix_rows(rows, module_name, filter_keyword)
        if filtered:
            sheets_data.append((module_name, sheet_base, filtered))

    if not sheets_data:
        for path in (output_path, os.path.splitext(output_path)[0] + ".pdf"):
            if os.path.exists(path):
                try:
                    os.remove(path)
                    print(f"Removed empty appendix output: {path}")
                except OSError as e:
                    print(f"Warning: could not remove {path}: {e}")
        print(f"No matching rows for Appendix {appendix_type}; skipping workbook.")
        return 0

    wb = openpyxl.Workbook()
    default = wb.active
    first = True
    sheets_created = 0

    for _module_name, sheet_base, filtered in sheets_data:
        title = unique_sheet_name(wb, sheet_base)
        if first:
            ws = default
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title)

        ws["A1"] = "ID"
        ws["B1"] = "Primary Text"
        ws["C1"] = "Artifact Type"
        ws["D1"] = "Change in Artifact Type"

        out_row = 2
        for row in filtered:
            ref_id = (row.get("REFERENCE_ID") or "").strip()
            primary = clean_text(row.get("URL1_title") or "")
            url = re.sub(r"[\r\n\t ]+", "", (row.get("URL1") or "").strip())
            artifact = (row.get("REQUIREMENT_TYPE") or "").strip()
            action = (row.get("LITERAL_NAME3") or "").strip()

            ws.cell(row=out_row, column=1, value=ref_id)

            b_cell = ws.cell(row=out_row, column=2)
            if url:
                b_cell.value = primary
                b_cell.hyperlink = url
            else:
                b_cell.value = primary

            if appendix_type == "B":
                ws.cell(row=out_row, column=3, value="PS Requirement")
            else:
                ws.cell(row=out_row, column=3, value=artifact)

            ws.cell(row=out_row, column=4, value=change_in_artifact_type(action))
            out_row += 1

        format_appendix_sheet(ws, out_row - 1, left_header)
        sheets_created += 1

    save_workbook(wb, output_path)
    return sheets_created


def ensure_cache_path(cache_path):
    if os.path.exists(cache_path):
        return cache_path
    legacy = os.path.join(script_dir(), "(OYB-SHD) Compliance Summary_AppB_AppC.xml")
    if os.path.exists(legacy):
        return legacy
    return cache_path


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Fill Requirement_Compliance_Summary.xlsx from XML and build "
            "Appendix B / Appendix C workbooks (and PDFs)."
        )
    )
    parser.add_argument("--xml", help="Force a local XML file (skip the live URL).")
    parser.add_argument(
        "--template",
        default=TEMPLATE_FILE,
        help="Path to Requirement_Compliance_Summary.xlsx template.",
    )
    parser.add_argument(
        "--summary",
        default=SUMMARY_FILE,
        help="Output path for the filled summary workbook.",
    )
    parser.add_argument("--cache", default=CACHE_FILE, help="Local XML cache path.")
    parser.add_argument(
        "--skip-appendices",
        action="store_true",
        help="Only update the summary sheet; do not write Appendix B/C.",
    )
    parser.add_argument("--skip-pdf", action="store_true", help="Do not export PDF via Excel.")
    return parser.parse_args()


def maybe_pdf(path, skip_pdf):
    if skip_pdf or not os.path.exists(path):
        return
    export_workbook_to_pdf(path)


def remove_stale_summary_pdf(summary_path):
    pdf_path = os.path.splitext(summary_path)[0] + ".pdf"
    if os.path.exists(pdf_path):
        try:
            os.remove(pdf_path)
            print(f"Removed stale summary PDF: {pdf_path}")
        except OSError as e:
            print(f"Warning: could not remove stale summary PDF {pdf_path}: {e}")


def main():
    args = parse_args()
    cache = ensure_cache_path(args.cache)
    xml_content = fetch_xml(URL, cache, xml_path=args.xml)
    print("\nParsing XML payload...")
    rows = parse_results(xml_content)

    if not os.path.exists(args.template):
        raise SystemExit(
            f"Summary template not found: {args.template}\n"
            "Provide Template/Requirement_Compliance_Summary.xlsx with columns A–O."
        )

    print(f"Loading summary template: {args.template}")
    wb = openpyxl.load_workbook(args.template)
    ws = wb.active
    if "Requirement_Compliance_Summary" in wb.sheetnames:
        ws = wb["Requirement_Compliance_Summary"]

    layout = locate_summary_layout(ws)
    by_name, seen_names = aggregate_counts(rows)
    catalog_names = fill_summary(ws, by_name, layout)
    save_workbook(wb, args.summary)
    remove_stale_summary_pdf(args.summary)

    unmatched = {
        name: n
        for name, n in seen_names.items()
        if name not in catalog_names and n > 0
    }
    if unmatched:
        print("XML module names not in summary catalog (counts ignored):")
        for name, n in sorted(unmatched.items(), key=lambda x: -x[1]):
            print(f"  {n:5d}  {name}")

    if args.skip_appendices:
        return

    modules_b = appendix_modules(ws, "B", layout)
    modules_c = appendix_modules(ws, "C", layout)
    print(f"\nBuilding Appendix B ({len(modules_b)} candidate modules)...")
    n_b = write_appendix_workbook(
        modules_b,
        rows,
        filter_keyword="Agreed",
        appendix_type="B",
        output_path=APPENDIX_B_FILE,
        left_header="Appendix B – List Of Changes Agreed By MTR In Artifact Types",
    )
    print(f"Appendix B sheets: {n_b}")
    if n_b:
        maybe_pdf(APPENDIX_B_FILE, args.skip_pdf)

    print(f"\nBuilding Appendix C ({len(modules_c)} candidate modules)...")
    n_c = write_appendix_workbook(
        modules_c,
        rows,
        filter_keyword="Proposed",
        appendix_type="C",
        output_path=APPENDIX_C_FILE,
        left_header="Appendix C – List Of Proposed Changes In Artifact Types",
    )
    print(f"Appendix C sheets: {n_c}")
    if n_c:
        maybe_pdf(APPENDIX_C_FILE, args.skip_pdf)


if __name__ == "__main__":
    main()
