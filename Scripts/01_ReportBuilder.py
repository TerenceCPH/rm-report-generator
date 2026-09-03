"""01 Report Builder: XML (live or cache) -> formatted Excel + PDF."""

import argparse
import os
import re

import openpyxl
from openpyxl.styles import Border, Font
from openpyxl.utils import get_column_letter

from env import (
    ALIGN_CENTER_WRAP,
    ALIGN_WRAP,
    FONT_NAME,
    GRAY_FILL,
    LINK_FONT_11,
    THICK_SIDE,
    THIN_BORDER,
    apply_print_setup,
    apply_thick_pair_borders,
    autofit_row_heights,
    export_workbook_to_pdf,
    fetch_xml,
    output_subdir,
    parse_results,
    save_workbook,
    xml_cache_path,
    xml_source_url,
)

URL = xml_source_url("report_builder")
CACHE_FILE = xml_cache_path("report_builder")
OUTPUT_DIR = output_subdir("01_ReportBuilder")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Appendix A – Report Builder.xlsx")

# Pair index -> (val_tag, id_tag, title_tag, url_tag, prefix_header, text_header)
VAL_PAIRS = [
    ("VAL", "REFERENCE_ID", "URL1_title", "URL1", "Doc Prefix", "Scope / IRS"),
    ("VAL2", "REFERENCE_ID1", "URL2_title", "URL2", "Satisfied by", "Design Deliverables (Level 1)"),
    ("VAL3", "REFERENCE_ID2", "URL3_title", "URL3", "Satisfied by", "Design Deliverables (Level 2)"),
    ("VAL4", "REFERENCE_ID3", "URL4_title", "URL4", "Verified by", "Design Deliverables (Level 3)"),
]


def build_schema(max_val_pairs):
    schema = []
    for i in range(max_val_pairs):
        val_tag, id_tag, title_tag, url_tag, prefix_header, text_header = VAL_PAIRS[i]
        schema.append(
            {"type": "normal", "tag": val_tag, "header": prefix_header, "width": 16, "hidden": False}
        )
        schema.append(
            {"type": "normal", "tag": id_tag, "header": "ID", "width": 12, "hidden": False}
        )
        schema.append(
            {
                "type": "hidden_title",
                "tag": title_tag,
                "id_tag": id_tag,
                "header": f"Raw Title {i + 1}",
                "width": 15,
                "hidden": True,
            }
        )
        schema.append(
            {
                "type": "hidden_url",
                "tag": url_tag,
                "header": f"Raw URL {i + 1}",
                "width": 15,
                "hidden": True,
            }
        )
        schema.append(
            {"type": "formula", "header": text_header, "width": 60, "hidden": False, "id_tag": id_tag}
        )
    return schema


def populate_sheet(ws, rows, schema):
    font_regular = Font(name=FONT_NAME, size=11)
    font_header = Font(name=FONT_NAME, size=11, bold=True, color="000000")

    # Column A is a 1-based row index; schema columns start at B.
    ws.append(["#"] + [col["header"] for col in schema])

    print("Populating data with hidden helper columns...")
    for row_idx, row_data in enumerate(rows, start=2):
        index_cell = ws.cell(row=row_idx, column=1)
        index_cell.value = row_idx - 1
        index_cell.font = font_regular
        index_cell.alignment = ALIGN_CENTER_WRAP

        for col_idx, col_def in enumerate(schema, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            kind = col_def["type"]

            if kind == "normal":
                cell.value = row_data.get(col_def["tag"], "")
                cell.font = font_regular
                if col_def["header"] in ("ID", "Doc Prefix", "Satisfied by", "Verified by"):
                    cell.alignment = ALIGN_CENTER_WRAP
                else:
                    cell.alignment = ALIGN_WRAP

            elif kind == "hidden_title":
                title_text = re.sub(r"[\r\n\t]+", " ", row_data.get(col_def["tag"], "")).strip()
                if len(title_text) > 255:
                    title_text = "Refer to Doors Next for full text."
                if not title_text and row_data.get(col_def["id_tag"], "").strip():
                    title_text = "Refer to Doors Next for full text."
                cell.value = title_text

            elif kind == "hidden_url":
                url_text = re.sub(r"[\r\n\t ]+", "", row_data.get(col_def["tag"], "").strip())
                cell.value = url_text

            elif kind == "formula":
                title_col = get_column_letter(col_idx - 2)
                url_col = get_column_letter(col_idx - 1)
                url_val = ws.cell(row=row_idx, column=col_idx - 1).value
                if url_val:
                    cell.value = f"=HYPERLINK({url_col}{row_idx}, {title_col}{row_idx})"
                    cell.font = LINK_FONT_11
                elif row_data.get(col_def["id_tag"], "").strip():
                    cell.value = f"={title_col}{row_idx}"
                    cell.font = font_regular
                else:
                    cell.value = ""
                    cell.font = font_regular
                cell.alignment = ALIGN_WRAP

    max_row = ws.max_row
    max_col = ws.max_column

    index_digits = max(len(str(max(0, max_row - 1))), len("#"))
    ws.column_dimensions["A"].width = index_digits + 2
    for col_idx, col_def in enumerate(schema, start=2):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_def["width"]
        if col_def.get("hidden"):
            ws.column_dimensions[letter].hidden = True

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if r == 1:
                cell.font = font_header
                cell.fill = GRAY_FILL
                cell.alignment = ALIGN_CENTER_WRAP
            cell.border = THIN_BORDER

    visible_col_indices = [
        idx for idx, col_def in enumerate(schema, start=2) if not col_def["hidden"]
    ]
    apply_thick_pair_borders(ws, visible_col_indices, max_row, chunk_size=3)

    # Header: thick top/bottom; keep pair-group verticals (A|B, F|G, K|L, P|Q).
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cb = cell.border
        cell.border = Border(
            left=cb.left,
            right=cb.right,
            top=THICK_SIDE,
            bottom=THICK_SIDE,
        )

    for r in range(1, max_row + 1):
        cell = ws.cell(row=r, column=1)
        cb = cell.border
        cell.border = Border(
            left=THICK_SIDE,
            right=THICK_SIDE,
            top=cb.top,
            bottom=THICK_SIDE if r == max_row else cb.bottom,
        )

    print("Calculating auto-fit row heights...")
    autofit_row_heights(ws, visible_col_indices, max_row, formula_title_offset=2)

    print("Configuring print layout...")
    apply_print_setup(
        ws,
        left_header='&"Arial,Regular"&11 Appendix A – Report Builder from DOORS Next',
    )
    ws.oddFooter.right.text = '&"Arial,Regular"&11 Page &[Page] of &[Pages]'
    ws.oddFooter.right.size = 11
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def parse_args():
    parser = argparse.ArgumentParser(
        description="Build Appendix A (Report Builder) Excel and PDF from live or cached XML."
    )
    parser.add_argument(
        "--max-val-pairs",
        type=int,
        default=3,
        choices=[1, 2, 3, 4],
        help="How many VAL pairs to display (VAL, VAL2, VAL3, VAL4). Default: 3.",
    )
    parser.add_argument("--xml", help="Force a local XML file (skip the live URL).")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Output .xlsx path.")
    parser.add_argument("--cache", default=CACHE_FILE, help="Local XML cache path.")
    parser.add_argument("--skip-pdf", action="store_true", help="Do not export PDF via Excel.")
    return parser.parse_args()


def main():
    args = parse_args()
    xml_content = fetch_xml(URL, args.cache, xml_path=args.xml)
    print("\nParsing XML payload...")
    rows = parse_results(xml_content)

    schema = build_schema(args.max_val_pairs)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUBMISSION"
    populate_sheet(ws, rows, schema)
    save_workbook(wb, args.output)
    if not args.skip_pdf:
        export_workbook_to_pdf(args.output)


if __name__ == "__main__":
    main()
