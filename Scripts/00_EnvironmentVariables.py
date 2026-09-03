"""Shared paths, XML fetch, parse, Excel helpers, and Excel-COM PDF export."""

import html
import math
import os
import xml.etree.ElementTree as ET

import requests
import urllib3
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

USERNAME = "terencechan"
PASSWORD = "MYJK23vm"

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPTS_DIR)
if not os.path.exists(os.path.join(ROOT_DIR, "Output")) and os.path.exists(
    os.path.join(SCRIPTS_DIR, "Output")
):
    # Backward-compatible fallback when the scripts still live at project root.
    ROOT_DIR = SCRIPTS_DIR

INPUT_DIR = os.path.join(ROOT_DIR, "Inputs")
TEMPLATE_DIR = os.path.join(INPUT_DIR, "Template")
MODULE_LIST_DIR = os.path.join(INPUT_DIR, "Module List")
DEFAULT_MODULE_LIST_CSV = os.path.join(MODULE_LIST_DIR, "RM Module List.csv")
SUBMISSION_DIR = os.path.join(INPUT_DIR, "Submission")
DOCUMENT_IMPORTER_DIR = os.path.join(INPUT_DIR, "DocumentImporter")
OUTPUT_DIR = os.path.join(ROOT_DIR, "Output")
XML_DIR = os.path.join(INPUT_DIR, "XML")

XML_SOURCES = {
    "module_list": {
        "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/408/dataservice?report=400",
        "cache": "(TUE) Module List.xml",
    },
    "compliance_summary": {
        "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/405/dataservice?report=397",
        "cache": "(TUE) Compliance Summary_AppB_AppC.xml",
    },
    "report_builder": {
        "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/403/dataservice?report=395",
        "cache": "(TUE) Report Builder.xml",
    },
    "requirements_checker": {
        "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/367/dataservice?report=359",
        "cache": "(TUE) Requirements Checker.xml",
    },
}

# XML_SOURCES = {
#     "module_list": {
#         "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/325/dataservice?report=317",
#         "cache": "(OYB-SHD) Module List.xml",
#     },
#     "compliance_summary": {
#         "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/145/dataservice?report=137",
#         "cache": "(OYB-SHD) Compliance Summary_AppB_AppC.xml",
#     },
#     "report_builder": {
#         "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/139/dataservice?report=131",
#         "cache": "(OYB-SHD) Report Builder.xml",
#     },
#     "requirements_checker": {
#         "url": "https://1263doorsapp.shuion.com.hk:9443/rs/query/369/dataservice?report=361",
#         "cache": "(OYB-SHD) Requirements Checker.xml",
#     },
# }

FONT_NAME = "Arial"
THIN_SIDE = Side(border_style="thin", color="000000")
THICK_SIDE = Side(border_style="thick", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="center")
ALIGN_CENTER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_LEFT_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
LINK_FONT_11 = Font(name=FONT_NAME, size=11, color="0000FF", underline="single")
LINK_FONT_12 = Font(name=FONT_NAME, size=12, color="0563C1", underline="single")


def script_dir():
    return ROOT_DIR


def output_subdir(*parts):
    path = os.path.join(OUTPUT_DIR, *parts)
    os.makedirs(path, exist_ok=True)
    return path


def xml_source_url(key):
    return XML_SOURCES[key]["url"]


def xml_cache_path(key):
    return os.path.join(XML_DIR, XML_SOURCES[key]["cache"])


def with_basic_auth_param(url):
    if "basicAuthenticationEnabled=" in url:
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}basicAuthenticationEnabled=true"


def clean_tag(full_tag):
    return full_tag.split("}")[-1]


def clean_text(text):
    if not text:
        return ""
    text = html.unescape(text)
    text = text.replace("&amp;", "&")
    text = text.replace("&#39;", "'")
    text = text.replace("&quot;", '"')
    text = text.replace("&lt;", "<")
    text = text.replace("&gt;", ">")
    text = text.replace("\xa0", " ")
    return text.strip()


def fetch_xml(url, cache_file, xml_path=None, timeout=15):
    """Return XML bytes from --xml, the live URL, or the local cache.

    Online is attempted first. A local --xml path skips the network.
    If the live request fails, the XML folder cache is used when present.
    """
    if xml_path:
        print(f"Using local XML file: {xml_path}")
        with open(xml_path, "rb") as f:
            return f.read()

    live_url = with_basic_auth_param(url)
    print(f"Attempting to fetch data from remote endpoint...\n  {live_url}")
    try:
        response = requests.get(
            live_url, auth=(USERNAME, PASSWORD), verify=False, timeout=timeout
        )
        response.raise_for_status()
        xml_content = response.content
        print("Remote fetch successful. Saving a copy to local XML cache...")
        try:
            cache_dir = os.path.dirname(cache_file)
            if cache_dir:
                os.makedirs(cache_dir, exist_ok=True)
            with open(cache_file, "wb") as f:
                f.write(xml_content)
        except IOError as e:
            print(f"Warning: Failed to update local XML cache file: {e}")
        return xml_content
    except (requests.exceptions.RequestException, OSError, Exception) as e:
        print(f"\n[!] Notice: Could not access the online source ({e}).")
        print("Checking for local fallback cache system...")
        if os.path.exists(cache_file):
            print(f"-> Local cache found ('{cache_file}'). Proceeding with cached backup.")
            with open(cache_file, "rb") as f:
                return f.read()
        print("Critical Error: Online source is unreachable and no local XML cache exists.")
        raise SystemExit(1)


def parse_results(xml_content):
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"Error parsing XML data structure: {e}")
        raise SystemExit(1)

    rows = []
    for result in root.findall(".//{*}result"):
        row = {}
        for child in result:
            row[clean_tag(child.tag)] = clean_text("".join(child.itertext()))
        rows.append(row)
    if not rows:
        print("No <result> elements found in the XML response.")
        raise SystemExit(1)
    print(f"Parsed {len(rows)} result rows.")
    return rows


def apply_print_setup(
    ws,
    left_header="",
    right_header="",
    orientation="landscape",
    fit_width=1,
    print_title_rows="1:1",
    left=0.25,
    right=0.25,
    top=0.75,
    bottom=0.75,
):
    if orientation == "landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    else:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins = PageMargins(
        left=left, right=right, top=top, bottom=bottom, header=0.3, footer=0.3
    )
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = False
    ws.page_setup.zoom = False
    if print_title_rows:
        ws.print_title_rows = print_title_rows
    if left_header:
        ws.oddHeader.left.text = left_header
    if right_header:
        ws.oddHeader.right.text = right_header


def autofit_row_heights(ws, visible_col_indices, max_row, formula_title_offset=2):
    """Estimate wrapped row heights from visible cell text (and hidden title for formulas)."""
    for r in range(1, max_row + 1):
        max_lines = 1
        for c in visible_col_indices:
            cell = ws.cell(row=r, column=c)
            val = str(cell.value or "")

            if cell.alignment.wrap_text is False:
                lines_in_cell = 1
            else:
                if val.startswith("=") and formula_title_offset:
                    val = str(ws.cell(row=r, column=c - formula_title_offset).value or "")
                col_letter = get_column_letter(c)
                col_width = ws.column_dimensions[col_letter].width or 15
                chars_per_line = max(1, int(col_width * 0.95))
                lines_in_cell = 0
                for line in val.split("\n"):
                    lines_in_cell += max(1, math.ceil(len(line) / chars_per_line))
            if lines_in_cell > max_lines:
                max_lines = lines_in_cell
        ws.row_dimensions[r].height = max_lines * 16.5 + 12


def apply_thick_pair_borders(ws, visible_col_indices, max_row, chunk_size=3):
    for i in range(0, len(visible_col_indices), chunk_size):
        chunk = visible_col_indices[i : i + chunk_size]
        start_col = chunk[0]
        end_col = chunk[-1]
        for r in range(1, max_row + 1):
            for c in chunk:
                cell = ws.cell(row=r, column=c)
                cb = cell.border
                left = THICK_SIDE if c == start_col else cb.left
                right = THICK_SIDE if c == end_col else cb.right
                top = THICK_SIDE if r == 1 else cb.top
                bottom = THICK_SIDE if r == max_row else cb.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def save_workbook(wb, output_file):
    try:
        out_dir = os.path.dirname(os.path.abspath(output_file))
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        wb.save(output_file)
        print(f"Success! Saved '{output_file}'.")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
        raise SystemExit(1)


def export_workbook_to_pdf(xlsx_path, pdf_path=None):
    """Export an .xlsx to PDF via Excel COM (all sheets). Returns the PDF path."""
    import pythoncom
    import win32com.client

    xlsx_path = os.path.abspath(xlsx_path)
    if pdf_path is None:
        pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
    else:
        pdf_path = os.path.abspath(pdf_path)

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    pdf_dir = os.path.dirname(pdf_path)
    if pdf_dir:
        os.makedirs(pdf_dir, exist_ok=True)

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xlsx_path, ReadOnly=True)
        # 0 = xlTypePDF
        wb.ExportAsFixedFormat(0, pdf_path)
        print(f"Success! Saved '{pdf_path}'.")
        return pdf_path
    finally:
        if wb is not None:
            wb.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()
