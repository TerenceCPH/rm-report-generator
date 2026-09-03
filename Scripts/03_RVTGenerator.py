"""03 RVT Generator: download DOORS module views, format Excel, export PDF."""

import argparse
import math
import os
import re
import time

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties
from playwright.sync_api import sync_playwright

from env import (
    DEFAULT_MODULE_LIST_CSV,
    MODULE_LIST_DIR,
    PASSWORD,
    ROOT_DIR,
    USERNAME,
    export_workbook_to_pdf,
    output_subdir,
)

DEFAULT_CSV_CANDIDATES = [
    os.path.join(ROOT_DIR, "Inputs", "Module List", "TUE_Module_List.csv"),
]

DESIRED_COLUMNS = [
    "id",
    "Doc Prefix",
    "KBP",
    "Primary Text",
    "Artifact Type",
    "Technical Verifiable",
    "Design Compliant Status",
    "Link:Satisfies (>)",
    "Link:Satisfied By (<)",
    "Link:To verify (>)",
    "Link:Verified By (<)",
    "EI (Change Request)",
    "Final Compliant status",
    "Remarks",
]

COLUMN_WIDTHS = {
    "Doc Prefix": 16,
    "Primary Text": 60,
    "Artifact Type": 12,
    "Technical Verifiable": 12,
    "Design Compliant Status": 12,
    "Link:Satisfies (>)": 32,
    "Link:Satisfied By (<)": 32,
    "Link:To verify (>)": 32,
    "Link:Verified By (<)": 32,
    "EI (Change Request)": 12,
    "Final Compliant status": 12,
    "Remarks": 32,
}

ROW_HEIGHT_COLUMNS = {
    "Primary Text": 55,
    "Artifact Type": 10,
    "Link:Verified By (<)": 28,
}

LEFT_ALIGN_HEADERS = {
    "Primary Text",
    "Link:Satisfies (>)",
    "Link:Satisfied By (<)",
    "Link:To verify (>)",
    "Link:Verified By (<)",
    "Remarks",
}

LOGIN_URL = (
    "https://1263doorsapp.shuion.com.hk:9443/rm/web#action=com.ibm.rdm.web.pages.showProjectDashboard&componentURI=https%3A%2F%2F1263doorsapp.shuion.com.hk%3A9443%2Frm%2Frm-projects%2F_npqXUCclEfCz_bTWSv7vxQ%2Fcomponents%2F_n2qRgCclEfCz_bTWSv7vxQ&viewType=all"
)

# LOGIN_URL = (
#     "https://1263doorsapp.shuion.com.hk:9443/rm/web#action="
#     "com.ibm.rdm.web.pages.showProjectDashboard&componentURI="
#     "https%3A%2F%2F1263doorsapp.shuion.com.hk%3A9443%2Frm%2Frm-projects"
#     "%2F_wt4CcCcoEfCz_bTWSv7vxQ%2Fcomponents%2F_wvhBRScoEfCz_bTWSv7vxQ"
#     "&vvc.configuration=https%3A%2F%2F1263doorsapp.shuion.com.hk%3A9443"
#     "%2Frm%2Fcm%2Fstream%2F_wvqLJicoEfCz_bTWSv7vxQ&viewType=all"
# )


def resolve_csv_path(explicit):
    if explicit:
        return explicit
    for path in DEFAULT_CSV_CANDIDATES:
        if os.path.exists(path):
            return path
    return DEFAULT_CSV_CANDIDATES[0]


def normalize_module_id(module_id):
    """Return a digit-only ModuleID string, or None if not a valid numeric id.

    Skips placeholders such as "To be submitted...", empty cells, and NaN.
    Accepts ints, whole-number floats (515961.0), and digit strings.
    """
    if module_id is None:
        return None
    if isinstance(module_id, float):
        if math.isnan(module_id) or not module_id.is_integer() or module_id <= 0:
            return None
        return str(int(module_id))
    if isinstance(module_id, int):
        return str(module_id) if module_id > 0 else None

    text = str(module_id).strip()
    if not text or text.casefold() in {"nan", "none", "null", ""}:
        return None
    if text.isdigit():
        return text
    try:
        value = float(text)
    except ValueError:
        return None
    if not value.is_integer() or value <= 0:
        return None
    return str(int(value))


def is_valid_module_id(module_id):
    """Return True when module_id is a non-empty numeric DOORS module id."""
    return normalize_module_id(module_id) is not None


def category_folder_name(index_id, category_name):
    """Build folder name like D1_Preliminaries from IndexID and CategoryName."""
    prefix = str(index_id).split("-", 1)[0].strip().upper()
    category = str(category_name or "Uncategorized").strip()
    return f"{prefix}_{category}"


def module_output_paths(output_dir, index_id, category_name, module_id):
    """Return category directory and xlsx path for one module row."""
    category_dir = os.path.join(
        output_dir, category_folder_name(index_id, category_name)
    )
    file_name = f"{index_id}_{module_id}.xlsx"
    save_path = os.path.join(category_dir, file_name)
    return category_dir, file_name, save_path


def parse_module_row(row):
    """Normalize CSV row fields from RM_Module_List.csv."""
    index_id = str(row["IndexID"]).strip()
    category_name = str(row.get("CategoryName", "")).strip()
    doc_name = str(row.get("DocumentName", "")).strip()
    doc_number = str(row.get("DocumentNumber", "")).strip()
    raw_module_id = row["ModuleID"]
    module_id = normalize_module_id(raw_module_id)
    if module_id is None:
        # Preserve raw text for skip logging (e.g. "To be submitted...").
        if raw_module_id is None or (
            isinstance(raw_module_id, float) and math.isnan(raw_module_id)
        ):
            module_id = ""
        else:
            module_id = str(raw_module_id).strip()
    if not doc_name:
        doc_name = doc_number or module_id
    return index_id, category_name, doc_name, doc_number, module_id


def _normalize_header(value):
    return str(value or "").strip().casefold()


def _header_column_map(ws):
    """Return normalized header -> 1-based column index for row 1."""
    mapping = {}
    for col in range(1, (ws.max_column or 1) + 1):
        key = _normalize_header(ws.cell(row=1, column=col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def _column_index_by_header(ws, header_name, header_map=None):
    """Resolve a header label to its 1-based column index."""
    key = _normalize_header(header_name)
    if header_map is not None:
        return header_map.get(key)
    for col in range(1, (ws.max_column or 1) + 1):
        if _normalize_header(ws.cell(row=1, column=col).value) == key:
            return col
    return None


def _estimate_wrapped_lines(cell_value, chars_per_line):
    lines = 0
    for paragraph in str(cell_value or "").split("\n"):
        lines += math.ceil(len(paragraph) / chars_per_line) if paragraph else 1
    return lines


def reorder_columns_by_header(ws):
    """Place DESIRED_COLUMNS in fixed order; drop any extra columns."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    source_by_header = {}
    for col in range(1, max_col + 1):
        key = _normalize_header(ws.cell(row=1, column=col).value)
        if key and key not in source_by_header:
            source_by_header[key] = col

    ordered = []
    for name in DESIRED_COLUMNS:
        src = source_by_header.get(_normalize_header(name))
        if src:
            ordered.append([ws.cell(row=r, column=src).value for r in range(1, max_row + 1)])
        else:
            ordered.append([name] + [None] * (max_row - 1))

    for dest, values in enumerate(ordered, start=1):
        for row, value in enumerate(values, start=1):
            ws.cell(row=row, column=dest).value = value

    leftover = max_col - len(DESIRED_COLUMNS)
    if leftover > 0:
        ws.delete_cols(len(DESIRED_COLUMNS) + 1, leftover)


def format_excel(file_path: str, doc_name: str):
    """
    Applies formatting to downloaded XLSX file:
    - Enables wrap_text on ALL columns.
    - Sets row height from wrapped-line estimates for Primary Text, Artifact Type,
      and Link:Verified By (<).
    - Deletes METADATA rows, then reorders columns by header names, before applying borders.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and "METADATA" in str(val).upper():
            start_row = max(1, r - 1)
            ws.delete_rows(start_row, ws.max_row - start_row + 1)
            break

    reorder_columns_by_header(ws)

    max_row = ws.max_row
    max_col = ws.max_column

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"

    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    ws.oddHeader.left.text = '&"Arial,Regular"&10 Appendix D — RVT Attachment'
    ws.oddHeader.right.text = f'&"Arial,Regular"&10 {doc_name}'

    header_cols = _header_column_map(ws)

    for header_name, width in COLUMN_WIDTHS.items():
        col_idx = _column_index_by_header(ws, header_name, header_cols)
        if col_idx is not None:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    font_regular = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    target_last_col = min(max_col, len(DESIRED_COLUMNS))
    left_align_cols = {
        col
        for name in LEFT_ALIGN_HEADERS
        if (col := _column_index_by_header(ws, name, header_cols)) is not None
    }

    for row in range(1, max_row + 1):
        is_header = row == 1

        if is_header:
            ws.row_dimensions[row].height = 25
        else:
            line_counts = []
            for header_name, chars_per_line in ROW_HEIGHT_COLUMNS.items():
                col_idx = _column_index_by_header(ws, header_name, header_cols)
                if col_idx is not None:
                    val = ws.cell(row=row, column=col_idx).value
                    line_counts.append(_estimate_wrapped_lines(val, chars_per_line))

            max_lines = max(line_counts) if line_counts else 1
            calculated_height = max(20, max_lines * 15)
            ws.row_dimensions[row].height = calculated_height

        for col in range(1, target_last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = font_bold if is_header else font_regular

            if is_header:
                cell.alignment = center_align
            else:
                if col in left_align_cols:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

    wb.save(file_path)


def maybe_pdf(xlsx_path, skip_pdf):
    if skip_pdf:
        return
    try:
        export_workbook_to_pdf(xlsx_path)
    except Exception as e:
        print(f"    [!] Error exporting PDF: {e}")


def automate_doors_export(csv_path, output_dir, skip_pdf=False):
    os.makedirs(output_dir, exist_ok=True)

    try:
        df = pd.read_csv(csv_path)
    except FileNotFoundError:
        print(f"Error: File not found at {csv_path}")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="msedge")
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()

        page.goto(LOGIN_URL)
        page.fill("#jazz_app_internal_LoginWidget_0_userId", USERNAME)
        page.fill("#jazz_app_internal_LoginWidget_0_password", PASSWORD)
        page.check('input[type="checkbox"]')
        page.click(".j-button-primary")
        page.wait_for_load_state("networkidle")

        artifacts_btn = page.locator('a[title="Artifacts"]').first
        if not artifacts_btn.evaluate('el => el.classList.contains("selected")'):
            artifacts_btn.click()
            page.wait_for_load_state("networkidle")

        artifacts_url = page.url

        process_total = 0
        for index, row in df.iterrows():
            index_id, category_name, doc_name, _doc_number, module_id = parse_module_row(
                row
            )
            if not is_valid_module_id(module_id):
                print(
                    f"[skip] Row {index + 1}/{len(df)} ({index_id}): "
                    f"no valid ModuleID ({module_id!r})."
                )
                continue

            process_total += 1
            category_dir, file_name, save_path = module_output_paths(
                output_dir, index_id, category_name, module_id
            )
            os.makedirs(category_dir, exist_ok=True)

            if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                print(
                    f"[⇄] Module {process_total}: {index_id} ({module_id}) already "
                    f"exists ({file_name}). Formatting & skipping download."
                )
                try:
                    format_excel(save_path, doc_name)
                    maybe_pdf(save_path, skip_pdf)
                except Exception as fe:
                    print(f"    [!] Error applying Excel formatting: {fe}")
                continue

            print(
                f"\n[+] Processing Module {process_total}: {index_id} "
                f"({module_id}) -> {category_folder_name(index_id, category_name)}"
            )

            try:
                search_input = page.locator(
                    'input.filterText[placeholder*="Type to filter"]'
                ).first
                search_input.wait_for(state="visible", timeout=15000)
                search_input.clear()
                search_input.fill(module_id)
                search_input.press("Enter")

                page.wait_for_load_state("networkidle")
                time.sleep(2)

                resource_link = page.locator(
                    f'a.jazz-ui-ResourceLink:has-text("{module_id}")'
                ).first
                if resource_link.count() == 0 or not resource_link.is_visible():
                    print(f"[-] Module {module_id} not found. Moving to next.")
                    continue

                resource_link.click()
                page.wait_for_load_state("networkidle")

                view_span = page.locator(
                    'span[title="1263 Full View v3: Shared, All modules"]'
                ).first
                view_span.wait_for(state="visible", timeout=15000)
                view_span.click(button="right")

                export_option = page.get_by_text("Export View...").last
                export_option.wait_for(state="visible", timeout=10000)
                export_option.click(force=True)

                xlsx_label = page.get_by_text("XLSX", exact=True).last
                xlsx_label.wait_for(state="visible", timeout=15000)
                xlsx_label.click(force=True)

                ok_btn = page.locator('button.j-button-primary:has-text("OK")').last
                ok_btn.wait_for(state="visible", timeout=10000)

                with page.expect_download(timeout=60000) as download_info:
                    ok_btn.click()

                download = download_info.value
                download.save_as(save_path)

                while not os.path.exists(save_path) or os.path.getsize(save_path) == 0:
                    time.sleep(0.5)

                print(f"[✓] Downloaded: {file_name}")

                format_excel(save_path, doc_name)
                print(f"[✓] Formatted: {file_name}")
                maybe_pdf(save_path, skip_pdf)

            except Exception as e:
                print(f"[!] Error processing Module {module_id}: {e}")
                print("    Recovering state and continuing to next module...")

            try:
                page.goto(artifacts_url)
                page.wait_for_load_state("networkidle")
                page.locator(
                    'input.filterText[placeholder*="Type to filter"]'
                ).first.wait_for(state="visible", timeout=15000)
                time.sleep(1)
            except Exception as nav_e:
                print(f"    [!] Failed to reset page navigation: {nav_e}")

        browser.close()
        print("\n[✓] All module downloads and formattings completed!")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Download RVT module views, format Excel, and export PDF."
    )
    parser.add_argument(
        "--csv",
        default=None,
        help=(
            "Path to RM_Module_List.csv "
            "(IndexID, CategoryName, DocumentName, DocumentNumber, ModuleID)."
        ),
    )
    parser.add_argument(
        "--output-dir",
        default=output_subdir("03_RVTGenerator"),
        help="Folder for per-module xlsx/pdf files.",
    )
    parser.add_argument("--skip-pdf", action="store_true", help="Do not export PDF via Excel.")
    return parser.parse_args()


def main():
    args = parse_args()
    csv_path = resolve_csv_path(args.csv)
    automate_doors_export(csv_path, args.output_dir, skip_pdf=args.skip_pdf)


if __name__ == "__main__":
    main()
