"""05 Requirements Checker: XML (live or cache) -> formatted Excel."""

import argparse
import os
import re

import openpyxl
from openpyxl.styles import Border, Font
from openpyxl.utils import get_column_letter

from env import (
    ALIGN_CENTER_WRAP,
    ALIGN_WRAP,
    FONT_NAME,
    GRAY_FILL,
    LINK_FONT_11,
    THICK_SIDE,
    THIN_BORDER,
    THIN_SIDE,
    apply_print_setup,
    autofit_row_heights,
    fetch_xml,
    output_subdir,
    parse_results,
    save_workbook,
    xml_cache_path,
    xml_source_url,
)

URL = xml_source_url("requirements_checker")
CACHE_FILE = xml_cache_path("requirements_checker")
OUTPUT_DIR = output_subdir("05_RequirementsChecker")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "RequirementsChecker.xlsx")

CENTER_HEADERS = {
    "ID",
    "Doc Prefix",
    "Satisfied by",
    "Requirement Type",
    "TV",
    "Compliance",
}

VAL_PAIRS = [
    {
        "val_tag": "VAL",
        "id_tag": "REFERENCE_ID",
        "title_tag": "URL1_title",
        "url_tag": "URL1",
        "prefix_header": "Doc Prefix",
        "text_header": "Scope / IRS",
        "attrs": [
            ("REQUIREMENT_TYPE", "Requirement Type", 20),
            ("LITERAL_NAME", "TV", 10),
            ("LITERAL_NAME1", "Compliance", 18),
        ],
    },
    {
        "val_tag": "VAL2",
        "id_tag": "REFERENCE_ID1",
        "title_tag": "URL2_title",
        "url_tag": "URL2",
        "prefix_header": "Satisfied by",
        "text_header": "Design Deliverables (Level 1)",
        "attrs": [
            ("REQUIREMENT_TYPE1", "Requirement Type", 20),
            ("LITERAL_NAME2", "TV", 10),
            ("LITERAL_NAME3", "Compliance", 18),
        ],
    },
    {
        "val_tag": "VAL4",
        "id_tag": "REFERENCE_ID2",
        "title_tag": "URL3_title",
        "url_tag": "URL3",
        "prefix_header": "Satisfied by",
        "text_header": "Design Deliverables (Level 2)",
        "attrs": [
            ("REQUIREMENT_TYPE2", "Requirement Type", 20),
        ],
    },
    {
        "val_tag": "VAL6",
        "id_tag": "REFERENCE_ID3",
        "title_tag": "URL4_title",
        "url_tag": "URL4",
        "prefix_header": "Satisfied by",
        "text_header": "Design Deliverables (Level 3)",
        "attrs": [
            ("REQUIREMENT_TYPE3", "Requirement Type", 20),
        ],
    },
    {
        "val_tag": None,
        "id_tag": "REFERENCE_ID4",
        "title_tag": "URL5_title",
        "url_tag": "URL5",
        "prefix_header": "Satisfied by",
        "text_header": "Design Deliverables (Level 4)",
        "attrs": [
            ("REQUIREMENT_TYPE4", "Requirement Type", 20),
        ],
    },
]


def build_schema(max_val_pairs):
    schema = []
    for i in range(max_val_pairs):
        pair = VAL_PAIRS[i]
        pair_index = i

        if pair["val_tag"]:
            schema.append(
                {
                    "type": "normal",
                    "tag": pair["val_tag"],
                    "header": pair["prefix_header"],
                    "width": 16,
                    "hidden": False,
                    "pair_index": pair_index,
                }
            )
        else:
            schema.append(
                {
                    "type": "empty",
                    "header": pair["prefix_header"],
                    "width": 16,
                    "hidden": False,
                    "pair_index": pair_index,
                }
            )

        schema.append(
            {
                "type": "normal",
                "tag": pair["id_tag"],
                "header": "ID",
                "width": 12,
                "hidden": False,
                "pair_index": pair_index,
            }
        )

        for tag, header, width in pair["attrs"]:
            schema.append(
                {
                    "type": "normal",
                    "tag": tag,
                    "header": header,
                    "width": width,
                    "hidden": False,
                    "pair_index": pair_index,
                }
            )

        schema.append(
            {
                "type": "hidden_title",
                "tag": pair["title_tag"],
                "id_tag": pair["id_tag"],
                "header": f"Raw Title {i + 1}",
                "width": 15,
                "hidden": True,
                "pair_index": pair_index,
            }
        )
        schema.append(
            {
                "type": "hidden_url",
                "tag": pair["url_tag"],
                "header": f"Raw URL {i + 1}",
                "width": 15,
                "hidden": True,
                "pair_index": pair_index,
            }
        )
        schema.append(
            {
                "type": "formula",
                "header": pair["text_header"],
                "width": 60,
                "hidden": False,
                "id_tag": pair["id_tag"],
                "pair_index": pair_index,
            }
        )
    return schema


def apply_pair_group_borders(ws, pair_col_groups, max_row):
    """Apply thick outer borders around each pair's visible column group."""
    for chunk in pair_col_groups:
        start_col = chunk[0]
        end_col = chunk[-1]
        for r in range(1, max_row + 1):
            for c in chunk:
                cell = ws.cell(row=r, column=c)
                cb = cell.border
                left = THICK_SIDE if c == start_col else cb.left
                right = THICK_SIDE if c == end_col else cb.right
                top = THICK_SIDE if r == 1 else cb.top
                bottom = THICK_SIDE if r == max_row else cb.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def visible_pair_groups(schema):
    pair_groups = {}
    for col_idx, col_def in enumerate(schema, start=2):
        if col_def.get("hidden"):
            continue
        pair_groups.setdefault(col_def["pair_index"], []).append(col_idx)
    return [pair_groups[i] for i in sorted(pair_groups)]


def populate_sheet(ws, rows, schema):
    font_regular = Font(name=FONT_NAME, size=11)
    font_header = Font(name=FONT_NAME, size=11, bold=True, color="000000")

    # Column A is a 1-based row index; schema columns start at B.
    ws.append(["#"] + [col["header"] for col in schema])

    print("Populating data with hidden helper columns...")
    for row_idx, row_data in enumerate(rows, start=2):
        index_cell = ws.cell(row=row_idx, column=1)
        index_cell.value = row_idx - 1
        index_cell.font = font_regular
        index_cell.alignment = ALIGN_CENTER_WRAP

        for col_idx, col_def in enumerate(schema, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            kind = col_def["type"]

            if kind == "normal":
                cell.value = row_data.get(col_def["tag"], "")
                cell.font = font_regular
                if col_def["header"] in CENTER_HEADERS:
                    cell.alignment = ALIGN_CENTER_WRAP
                else:
                    cell.alignment = ALIGN_WRAP

            elif kind == "empty":
                cell.value = ""
                cell.font = font_regular
                cell.alignment = ALIGN_CENTER_WRAP

            elif kind == "hidden_title":
                title_text = re.sub(r"[\r\n\t]+", " ", row_data.get(col_def["tag"], "")).strip()
                if len(title_text) > 255:
                    title_text = "Refer to Doors Next for full text."
                if not title_text and row_data.get(col_def["id_tag"], "").strip():
                    title_text = "Refer to Doors Next for full text."
                cell.value = title_text

            elif kind == "hidden_url":
                url_text = re.sub(r"[\r\n\t ]+", "", row_data.get(col_def["tag"], "").strip())
                cell.value = url_text

            elif kind == "formula":
                title_col = get_column_letter(col_idx - 2)
                url_col = get_column_letter(col_idx - 1)
                url_val = ws.cell(row=row_idx, column=col_idx - 1).value
                if url_val:
                    cell.value = f"=HYPERLINK({url_col}{row_idx}, {title_col}{row_idx})"
                    cell.font = LINK_FONT_11
                elif row_data.get(col_def["id_tag"], "").strip():
                    cell.value = f"={title_col}{row_idx}"
                    cell.font = font_regular
                else:
                    cell.value = ""
                    cell.font = font_regular
                cell.alignment = ALIGN_WRAP

    max_row = ws.max_row
    max_col = ws.max_column

    index_digits = max(len(str(max(0, max_row - 1))), len("#"))
    ws.column_dimensions["A"].width = index_digits + 2
    for col_idx, col_def in enumerate(schema, start=2):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_def["width"]
        if col_def.get("hidden"):
            ws.column_dimensions[letter].hidden = True

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if r == 1:
                cell.font = font_header
                cell.fill = GRAY_FILL
                cell.alignment = ALIGN_CENTER_WRAP
            cell.border = THIN_BORDER

    visible_col_indices = [
        idx for idx, col_def in enumerate(schema, start=2) if not col_def["hidden"]
    ]
    apply_pair_group_borders(ws, visible_pair_groups(schema), max_row)

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.border = Border(
            left=THICK_SIDE if c == 1 else THIN_SIDE,
            right=THICK_SIDE if c == max_col else THIN_SIDE,
            top=THICK_SIDE,
            bottom=THICK_SIDE,
        )

    for r in range(1, max_row + 1):
        cell = ws.cell(row=r, column=1)
        cb = cell.border
        cell.border = Border(
            left=THICK_SIDE,
            right=cb.right,
            top=cb.top,
            bottom=THICK_SIDE if r == max_row else cb.bottom,
        )

    print("Calculating auto-fit row heights...")
    autofit_row_heights(ws, visible_col_indices, max_row, formula_title_offset=2)

    print("Configuring print layout...")
    apply_print_setup(
        ws,
        left_header='&"Arial,Regular"&11 Requirements Checker from DOORS Next',
    )
    ws.oddFooter.right.text = '&"Arial,Regular"&11 Page &[Page] of &[Pages]'
    ws.oddFooter.right.size = 11
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def parse_args():
    parser = argparse.ArgumentParser(
        description="Build Requirements Checker Excel from live or cached XML."
    )
    parser.add_argument(
        "--max-val-pairs",
        type=int,
        default=3,
        choices=[1, 2, 3, 4, 5],
        help="How many VAL pairs to display (Scope + Levels 1–4). Default: 3.",
    )
    parser.add_argument("--xml", help="Force a local XML file (skip the live URL).")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Output .xlsx path.")
    parser.add_argument("--cache", default=CACHE_FILE, help="Local XML cache path.")
    return parser.parse_args()


def main():
    args = parse_args()
    xml_content = fetch_xml(URL, args.cache, xml_path=args.xml)
    print("\nParsing XML payload...")
    rows = parse_results(xml_content)

    schema = build_schema(args.max_val_pairs)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUBMISSION"
    populate_sheet(ws, rows, schema)
    save_workbook(wb, args.output)


if __name__ == "__main__":
    main()
